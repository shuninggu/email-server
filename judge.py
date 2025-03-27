import openai
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import os
import csv  # 放在文件开头


########################
# 第 1 部分：LLM 评估函数
########################

def judge_replies(reply_a: str, reply_b: str) -> str:
    """
    对比 reply_a, reply_b，调用评估用的 LLM 返回:
    - "model1"  (表示 A 更好)
    - "model2"  (表示 B 更好)
    - "tie"     (表示 差不多)
    """

    # 这里替换成你自己的 API Key 或者调用私有 LLM 的接口
    # openai.api_key = "YOUR_OPENAI_API_KEY"
    openai.api_key = os.environ.get("OPENAI_API_KEY")
    if openai.api_key is None:
        raise ValueError("❌ Environment variable 'OPENAI_API_KEY' not found. Please export it in your terminal.")


    # 给评估模型的提示词，可根据需求改动
    system_prompt = """You are an unbiased judge. 
You will be provided with two email replies (Reply A and Reply B). 
You must compare them in terms of clarity, correctness, completeness, politeness, and overall helpfulness. 
Output one single token among these:
- "model1": if Reply A is clearly better,
- "model2": if Reply B is clearly better,
- "tie": if they are about the same quality.
Do not explain your reasoning. Just return one word exactly: model1, model2, or tie.
"""
    user_prompt = f"""
Reply A:
{reply_a}

Reply B:
{reply_b}

Which reply is better?
"""
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",  # 或 "gpt-3.5-turbo" / 你自己部署的模型
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user",   "content": user_prompt},
            ],
            temperature=0.0
        )
        content = response["choices"][0]["message"]["content"].strip().lower()
        if content in ["model1", "model2", "tie"]:
            return content
        else:
            return "tie"
    except Exception as e:
        print("Error calling LLM:", e)
        # 如果出错就默认设为 tie，或根据需要改成其它逻辑
        return "tie"

########################
# 第 2 部分：Excel 数据读取 & 两两对比
########################

# def compare_two_columns(
#     workbook_path: str,
#     col_a: int,
#     col_b: int,
#     skip_header: bool = True,
#     judge_func=judge_replies,
# ) -> dict:
#     """
#     从 Excel 中读取两列数据 (列号从1开始)，
#     调用 judge_func() 进行对比，统计 modelA胜 / modelB胜 / tie 的次数。
    
#     返回形如: {"model1": x, "model2": y, "tie": z, "total": n}
#     """
#     wb = load_workbook(workbook_path)
#     sheet = wb.active
#     rows = list(sheet.iter_rows(values_only=True))

#     if skip_header:
#         rows = rows[1:]  # 跳过表头行

#     score_model1 = 0
#     score_model2 = 0
#     score_tie = 0
#     total_count = 0

#     for row in rows:
#         # row 是一个元组，每个元素对应一列的值
#         # Excel 中列号 col_a, col_b 是从1开始，但 Python 下标从0开始，所以要 col_a-1
#         reply_a = row[col_a - 1] if len(row) >= col_a else None
#         reply_b = row[col_b - 1] if len(row) >= col_b else None
#         # 如果其中一个为空，就跳过
#         if not reply_a or not reply_b:
#             continue
#         print(f"\n🔍 Comparing pair #{total_count + 1}")
#         print(f"--- Reply A (Model1):\n{reply_a}")
#         print(f"--- Reply B (Model2):\n{reply_b}")
#         total_count += 1
#         result = judge_func(reply_a, reply_b)
#         if result == "model1":
#             score_model1 += 1
#         elif result == "model2":
#             score_model2 += 1
#         else:
#             score_tie += 1

#     return {
#         "model1": score_model1,
#         "model2": score_model2,
#         "tie": score_tie,
#         "total": total_count,
#     }

def compare_two_columns(
    workbook_path: str,
    col_a: int,
    col_b: int,
    skip_header: bool = True,
    judge_func=judge_replies,
    log_path: str = None,
    label: str = None
) -> dict:
    """
    从 Excel 中读取两列数据 (列号从1开始)，
    调用 judge_func() 进行对比，统计 modelA胜 / modelB胜 / tie 的次数。
    可选：将每一条对比写入 log 文件。
    """
    wb = load_workbook(workbook_path)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if skip_header:
        rows = rows[1:]

    score_model1 = 0
    score_model2 = 0
    score_tie = 0
    total_count = 0

    logs = []

    for idx, row in enumerate(rows):
        reply_a = row[col_a - 1] if len(row) >= col_a else None
        reply_b = row[col_b - 1] if len(row) >= col_b else None
        if not reply_a or not reply_b:
            continue
        total_count += 1
        result = judge_func(reply_a, reply_b)
        if result == "model1":
            score_model1 += 1
        elif result == "model2":
            score_model2 += 1
        else:
            score_tie += 1

        # <<< 添加日志 >>>
        logs.append({
            "Index": idx + 2 if skip_header else idx + 1,
            "Reply A": reply_a,
            "Reply B": reply_b,
            "Result": result
        })

    # <<< 写入日志到 CSV >>>
    if log_path and label:
        filename = f"{log_path}/comparison_log_{label.replace(' ', '_')}.csv"
        with open(filename, mode="w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["Index", "Reply A", "Reply B", "Result"])
            writer.writeheader()
            writer.writerows(logs)
        print(f"🔍 Log saved to {filename}")

    return {
        "model1": score_model1,
        "model2": score_model2,
        "tie": score_tie,
        "total": total_count,
    }
    
########################
# 第 3 部分：可视化 - 画堆叠柱状图
########################

def create_stacked_bar_chart(results_list: list, title: str = "Human Evaluation"):
    """
    给定若干组对比结果，画出堆叠柱状图，每组显示 model1_win, tie, model2_win 三段。
    
    results_list 形如：
    [
      {
        "label": "ModelA vs ModelB",
        "model1": 10,
        "model2": 5,
        "tie": 2,
        "total": 17
      },
      {
        "label": "ModelC vs ModelD",
        "model1": 8,
        "model2": 8,
        "tie": 4,
        "total": 20
      },
      ...
    ]
    """
    labels = [r["label"] for r in results_list]
    # 计算百分比
    model1_pct = []
    tie_pct = []
    model2_pct = []
    for r in results_list:
        total = r["total"]
        if total == 0:
            model1_pct.append(0)
            tie_pct.append(0)
            model2_pct.append(0)
        else:
            model1_pct.append(r["model1"] / total * 100)
            tie_pct.append(r["tie"] / total * 100)
            model2_pct.append(r["model2"] / total * 100)

    x = range(len(labels))  # x 轴每一组位置
    fig, ax = plt.subplots(figsize=(8, 4))

    # 画三段堆叠: model1_win 在最底层，tie 在上面，model2_win 在最上面
    bar1 = ax.barh(x, model1_pct, color='teal', label='Win (Model1)')
    bar2 = ax.barh(x, tie_pct, left=model1_pct, color='gray', label='Tie')
    # 这里的 left= 需要是 model1_pct + tie_pct 的累积
    # 但 barh() 只认具体数值而非列表间加法，所以得先把它们加起来
    left_for_model2 = [m1 + t for m1, t in zip(model1_pct, tie_pct)]
    bar3 = ax.barh(x, model2_pct, left=left_for_model2, color='red', label='Win (Model2)')

    # 在每段条形上加上数值
    def add_labels(bar_obj, base_list):
        for rect, base_val in zip(bar_obj, base_list):
            width = rect.get_width()
            if width > 0:
                # 显示在区段中间
                x_pos = base_val + width / 2
                y_pos = rect.get_y() + rect.get_height() / 2
                ax.text(x_pos, y_pos, f"{width:.1f}%", ha='center', va='center', color='white', fontsize=9)
    
    add_labels(bar1, [0]*len(model1_pct))
    add_labels(bar2, model1_pct)
    add_labels(bar3, left_for_model2)

    ax.set_yticks(x)
    ax.set_yticklabels(labels)
    ax.set_xlabel("% win rate")
    ax.set_xlim([0, 100])
    ax.set_title(title)
    ax.invert_yaxis()  # 让最上面一组在图上端，可根据审美决定是否 invert
    ax.legend(loc="lower right")
    plt.tight_layout()
    plt.show()

########################
# 第 4 部分：示例主函数 - 自定义需要比较的「列对」
########################

def main():
    workbook_path = "results.xlsx"
    
    # 这里定义若干要对比的“列对”，每个元素是 (name, colA, colB)
    # 注：列号从 1 开始数，如：1 = A列, 2 = B列, 3 = C列, 4 = D列, ...
    pairs_to_compare = [
        ("local Llama3.2-3b vs GPT-4", 2, 5),  # D列 vs E列 (示例)
        ("someModel vs anotherModel", 2, 3),  # B列 vs C列 (举例)
        # 你也可以添加更多，如 ("ModelX vs ModelY", 6, 7) ...
    ]
    
    results_for_chart = []
    
    for pair_label, colA, colB in pairs_to_compare:
        # comparison = compare_two_columns(
        #     workbook_path=workbook_path,
        #     col_a=colA,
        #     col_b=colB,
        #     skip_header=True,
        #     judge_func=judge_replies
        # )
        
        comparison = compare_two_columns(
            workbook_path=workbook_path,
            col_a=colA,
            col_b=colB,
            skip_header=True,
            judge_func=judge_replies,
            log_path="logs",              # 你想存日志的文件夹
            label=pair_label              # 用来命名日志文件
        )
        
        # 把对比结果汇总到一个字典里，用于后面绘图
        results_for_chart.append({
            "label": pair_label,
            "model1": comparison["model1"],
            "model2": comparison["model2"],
            "tie": comparison["tie"],
            "total": comparison["total"],
        })
    
    # 绘制堆叠柱状图
    create_stacked_bar_chart(results_for_chart, title="Pairwise Comparison Results")

if __name__ == "__main__":
    main()

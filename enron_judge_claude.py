import os
import csv
import time
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import anthropic  # Claude API

########################
# Claude-based evaluation function
########################

def judge_replies(reply_a: str, reply_b: str) -> str:
    """
    Compare two replies using Claude API.
    Return one of:
    - "model1": if Reply A is better
    - "model2": if Reply B is better
    - "tie": if both are about the same
    """
    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])

#     system_prompt = """You are an unbiased judge.
# You will be provided with two email replies (Reply A and Reply B).
# Evaluate them based on:
# 1. Naturalness: The reply should sound like something a human would naturally write.
# 2. Succinctness: It should be concise and to the point, avoiding unnecessary repetition.

# Only reply with one of the following words:
# - model1
# - model2
# - tie
# Do not explain. Do not include anything else.
# """

    system_prompt = """
    You are an unbiased judge.

You will be shown two email replies: Reply A and Reply B. Your task is to decide which one is better based on the following criteria:

1. Naturalness: Does the reply sound like it was written by a human? Is it appropriate in tone and style?
2. Succinctness: Is the reply clear, concise, and free of unnecessary repetition?

You must choose the better reply unless they are truly indistinguishable in both criteria. Use your best judgment — do not default to “tie” unless there is no meaningful difference.

Respond with exactly one word, and nothing else:
- model1  (if Reply A is better)
- model2  (if Reply B is better)
- tie     (only if they are *truly* equal in both criteria)
"""

    user_prompt = f"""
Reply A:
{reply_a}

Reply B:
{reply_b}

Which reply is better?
"""

    try:
        response = client.messages.create(
            model="claude-3-sonnet-20240229",  # Change to sonnet/opus if desired
            max_tokens=1,
            temperature=0,
            system=system_prompt,
            messages=[
                {"role": "user", "content": user_prompt}
            ]
        )
        result = response.content[0].text.strip().lower()
        if result in ["model1", "model2", "tie"]:
            return result
        else:
            return "tie"
    except Exception as e:
        print("❌ Claude API error:", e)
        return "tie"

########################
# Compare two columns from Excel
########################

def compare_two_columns(
    workbook_path: str,
    col_a: int,
    col_b: int,
    skip_header: bool = True,
    judge_func=judge_replies,
    log_path: str = None,
    label: str = None
) -> dict:
    """
    Compare two Excel columns (1-indexed). Log each comparison result.
    Return win/tie counts.
    """
    wb = load_workbook(workbook_path)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if skip_header:
        rows = rows[1:]

    score_modelA = 0
    score_modelB = 0
    score_tie = 0
    total_count = 0
    logs = []

    for idx, row in enumerate(rows):
        reply_a = row[col_a - 1] if len(row) >= col_a else None
        reply_b = row[col_b - 1] if len(row) >= col_b else None
        if not reply_a or not reply_b:
            continue

        total_count += 1
        start_time = time.time()
        result_raw = judge_func(reply_a, reply_b)
        elapsed = time.time() - start_time

        if result_raw == "model1":
            result_str = "modelA wins"
            score_modelA += 1
        elif result_raw == "model2":
            result_str = "modelB wins"
            score_modelB += 1
        else:
            result_str = "tie"
            score_tie += 1

        logs.append({
            "Index": idx + 2 if skip_header else idx + 1,
            "Model A": reply_a,
            "Model B": reply_b,
            "Compare Result": result_str,
            "Time (sec)": f"{elapsed:.2f}"
        })

    if log_path and label:
        os.makedirs(log_path, exist_ok=True)
        filename = f"{log_path}/comparison_log_{label.replace(' ', '_')}.csv"
        with open(filename, mode="w", encoding="utf-8-sig", newline="") as f:
            fieldnames = ["Index", "Model A", "Model B", "Compare Result", "Time (sec)"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(logs)
        print(f"✅ Log saved to: {filename}")

    return {
        "modelA": score_modelA,
        "modelB": score_modelB,
        "tie": score_tie,
        "total": total_count,
    }

########################
# Horizontal stacked bar chart
########################

def create_stacked_bar_chart(results_list: list, title: str = "Human Evaluation"):
    """
    Plot win/tie percentages for each model pair comparison.
    """
    labels = [r["label"] for r in results_list]
    modelA_pct, tie_pct, modelB_pct = [], [], []

    for r in results_list:
        total = r["total"]
        if total == 0:
            modelA_pct.append(0)
            tie_pct.append(0)
            modelB_pct.append(0)
        else:
            modelA_pct.append(r["modelA"] / total * 100)
            tie_pct.append(r["tie"] / total * 100)
            modelB_pct.append(r["modelB"] / total * 100)

    x_positions = range(len(labels))
    fig, ax = plt.subplots(figsize=(8, 4))

    barA = ax.barh(x_positions, modelA_pct, label='Win (ModelA)')
    barT = ax.barh(x_positions, tie_pct, left=modelA_pct, label='Tie')
    left_for_B = [a + t for a, t in zip(modelA_pct, tie_pct)]
    barB = ax.barh(x_positions, modelB_pct, left=left_for_B, label='Win (ModelB)')

    def add_labels(bar_obj, base_list):
        for rect, base_val in zip(bar_obj, base_list):
            width = rect.get_width()
            if width > 0:
                x_pos = base_val + width / 2
                y_pos = rect.get_y() + rect.get_height() / 2
                ax.text(x_pos, y_pos, f"{width:.1f}%", ha='center', va='center', color='white', fontsize=9)

    add_labels(barA, [0]*len(modelA_pct))
    add_labels(barT, modelA_pct)
    add_labels(barB, left_for_B)

    ax.set_yticks(list(x_positions))
    ax.set_yticklabels(labels)
    ax.set_xlabel("% win rate")
    ax.set_xlim([0, 100])
    ax.set_title(title)
    ax.invert_yaxis()
    ax.legend(loc="lower right")
    plt.tight_layout()
    plt.show()

########################
# Main entry point
########################

def main():
    workbook_path = "enron_output.xlsx"

    # Define (label, column A, column B)
    pairs_to_compare = [
        ("gemma3:1b vs GPT-4", 2, 7),
        ("gemma:2b vs GPT-4", 3, 7),
        ("llama3.2:3b vs GPT-4", 4, 7),
        ("mistral vs GPT-4", 5, 7),
    ]

    results_for_chart = []

    for pair_label, colA, colB in pairs_to_compare:
        comparison = compare_two_columns(
            workbook_path=workbook_path,
            col_a=colA,
            col_b=colB,
            skip_header=True,
            judge_func=judge_replies,
            log_path="logs_claude_2",
            label=pair_label
        )

        results_for_chart.append({
            "label": pair_label,
            "modelA": comparison["modelA"],
            "modelB": comparison["modelB"],
            "tie": comparison["tie"],
            "total": comparison["total"],
        })

    create_stacked_bar_chart(results_for_chart, title="Claude-Based Email Reply Evaluation")

if __name__ == "__main__":
    main()

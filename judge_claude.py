import os
import csv
import time
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import anthropic  # Claude API

########################
# Claude-based evaluation function
########################

def judge_replies(original_email: str, reply_a: str, reply_b: str) -> str:
    """
    Compare two replies using Claude API, also providing the original email for context.
    Return one of:
    - "model1": if Reply A is better
    - "model2": if Reply B is better
    - "tie": if both are about the same
    """
    client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])

    # 系统提示词
    system_prompt = """
    You are an unbiased judge.

    You will be shown an original email, and two different replies (Reply A and Reply B).
    Your task is to decide which reply is better based on the following criteria:

    1. Naturalness: Does the reply sound like it was written by a human? Is it appropriate in tone and style?
    2. Succinctness: Is the reply clear, concise, and free of unnecessary repetition?

    Respond with exactly one of these options and nothing else:
    - model1
    - model2
    - tie
    """

    # 用户提示词，包含 Original Email、Reply A、Reply B
    user_prompt = f"""
Original Email:
{original_email}

Reply A:
{reply_a}

Reply B:
{reply_b}

Which reply is better? Answer with exactly one word: model1, model2, or tie.
"""

    try:
        response = client.messages.create(
            # model="claude-3-sonnet-20240229",  # Use the latest available model
            model="claude-3-7-sonnet-latest",  # Use the latest available model
            max_tokens=10,  # 增加一点保证能捕获完整答案
            temperature=0,
            system=system_prompt,
            messages=[
                {"role": "user", "content": user_prompt}
            ]
        )

        # 注意这里的解析方式，根据你的实际数据结构进行调整
        # 假设 response.content[0].text 为 Claude 的文本输出
        result = response.content[0].text.strip().lower()

        if "model1" in result:
            return "model1"
        elif "model2" in result:
            return "model2"
        elif "tie" in result:
            return "tie"
        else:
            print(f"⚠️ Unexpected response: {result}")
            return "tie"
    except Exception as e:
        print("❌ Claude API error:", e)
        return "tie"

########################
# Compare two columns from Excel
########################

def compare_two_columns(
    workbook_path: str,
    col_a: int,
    col_b: int,
    skip_header: bool = True,
    judge_func=judge_replies,
    log_path: str = None,
    label: str = None
) -> dict:
    """
    Compare two Excel columns (1-indexed). 
    Also includes column 1 (original email) as context in judge_func.
    Log each comparison result, return overall stats.
    """
    wb = load_workbook(workbook_path)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if skip_header:
        rows = rows[1:]

    score_modelA = 0
    score_modelB = 0
    score_tie = 0
    total_count = 0
    logs = []

    for idx, row in enumerate(rows):
        # row[0]假设就是第一列：Original Email
        original_email = row[0] if len(row) >= 1 and row[0] else ""
        reply_a = row[col_a - 1] if len(row) >= col_a and row[col_a - 1] else ""
        reply_b = row[col_b - 1] if len(row) >= col_b and row[col_b - 1] else ""

        # 如果 reply_a 或 reply_b 为空，则跳过此行
        if not reply_a or not reply_b:
            continue

        total_count += 1
        start_time = time.time()

        # 将 original_email 一并传入
        result_raw = judge_func(original_email, reply_a, reply_b)
        elapsed = time.time() - start_time

        if result_raw == "model1":
            result_str = "modelA wins"
            score_modelA += 1
        elif result_raw == "model2":
            result_str = "modelB wins"
            score_modelB += 1
        else:
            result_str = "tie"
            score_tie += 1

        logs.append({
            "Index": idx + 2 if skip_header else idx + 1,
            "Original Email": original_email,
            "Model A": reply_a,
            "Model B": reply_b,
            "Compare Result": result_str,
            "Time (sec)": f"{elapsed:.2f}"
        })

    if log_path and label:
        os.makedirs(log_path, exist_ok=True)
        filename = f"{log_path}/comparison_log_{label.replace(' ', '_')}.csv"
        with open(filename, mode="w", encoding="utf-8-sig", newline="") as f:
            fieldnames = ["Index", "Original Email", "Model A", "Model B", "Compare Result", "Time (sec)"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(logs)
        print(f"✅ Log saved to: {filename}")

    return {
        "modelA": score_modelA,
        "modelB": score_modelB,
        "tie": score_tie,
        "total": total_count,
    }

########################
# Horizontal stacked bar chart
########################

def create_stacked_bar_chart(results_list: list, title: str = "Model Evaluation", save_path: str = None):
    """
    Plot win/tie percentages for each model pair comparison.
    Optionally save the figure to `save_path`.
    """
    labels = [r["label"] for r in results_list]
    modelA_pct, tie_pct, modelB_pct = [], [], []

    for r in results_list:
        total = r["total"]
        if total == 0:
            modelA_pct.append(0)
            tie_pct.append(0)
            modelB_pct.append(0)
        else:
            modelA_pct.append(r["modelA"] / total * 100)
            tie_pct.append(r["tie"] / total * 100)
            modelB_pct.append(r["modelB"] / total * 100)

    x_positions = range(len(labels))
    fig, ax = plt.subplots(figsize=(8, 4))

    barA = ax.barh(x_positions, modelA_pct, label='Win (ModelA)')
    barT = ax.barh(x_positions, tie_pct, left=modelA_pct, label='Tie')
    left_for_B = [a + t for a, t in zip(modelA_pct, tie_pct)]
    barB = ax.barh(x_positions, modelB_pct, left=left_for_B, label='Win (ModelB)')

    def add_labels(bar_obj, base_list):
        for rect, base_val in zip(bar_obj, base_list):
            width = rect.get_width()
            if width > 0:
                x_pos = base_val + width / 2
                y_pos = rect.get_y() + rect.get_height() / 2
                ax.text(x_pos, y_pos, f"{width:.1f}%", ha='center', va='center', color='white', fontsize=9)

    add_labels(barA, [0]*len(modelA_pct))
    add_labels(barT, modelA_pct)
    add_labels(barB, left_for_B)

    ax.set_yticks(list(x_positions))
    ax.set_yticklabels(labels)
    ax.set_xlabel("% Win Rate")
    ax.set_xlim([0, 100])
    ax.set_title(title)
    ax.invert_yaxis()
    ax.legend(loc="lower right")
    plt.tight_layout()

    # 保存图片到指定路径
    if save_path:
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        fig.savefig(save_path)
        print(f"✅ Chart saved to: {save_path}")

    # 显示图形（如果不需要在屏幕上显示，可以注释掉）
    plt.show()

########################
# Main entry point
########################

def main():
    workbook_path = "all/replies.xlsx"

    # Define (label, column A, column B)
    # 注意：表格第一列是 original email (列索引=1)；第二列第三列等为不同模型回复
    # 此处只是演示：我们要比较 colA vs colB。你可以根据实际表格中的列号来修改。
    # pairs_to_compare = [
    #     ("gemma3:1b vs GPT-4o", 2, 6),
    #     ("gemma:2b vs GPT-4o", 3, 6),
    #     ("llama3.2:3b vs GPT-4o", 4, 6),
    #     ("mistral vs GPT-4o", 5, 6),
    # ]
    pairs_to_compare = [
        ("llama3.2:3b_masked vs llama3.2:3b", 11, 4),
        ("llama3.2:3b_masked vs GPT-4o", 11, 6),
    ]

    results_for_chart = []

    for pair_label, colA, colB in pairs_to_compare:
        comparison = compare_two_columns(
            workbook_path=workbook_path,
            col_a=colA,
            col_b=colB,
            skip_header=True,
            judge_func=judge_replies,
            # log_path="enron/logs_4local_gpt4o_3",
            log_path="all/claude_judge_llama3.2_3b",
            label=pair_label
        )

        results_for_chart.append({
            "label": pair_label,
            "modelA": comparison["modelA"],
            "modelB": comparison["modelB"],
            "tie": comparison["tie"],
            "total": comparison["total"],
        })

    # 图片保存路径
    chart_save_path = os.path.join("enron", "logs_4local_gpt4o_3", "evaluation_chart.png")
    create_stacked_bar_chart(
        results_for_chart,
        title="Claude-Based Email Reply Evaluation",
        save_path=chart_save_path
    )

if __name__ == "__main__":
    main()

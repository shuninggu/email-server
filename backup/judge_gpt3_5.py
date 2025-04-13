import os
import csv
import time
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# 如果你使用 openai 包，则在这里:
import openai
openai.api_key = os.environ.get("OPENAI_API_KEY")

# 如果你使用的是私有的 LLM 客户端（例如 class OpenAI(api_key=...) ），则保持如下引入:
from openai import OpenAI


########################
# 第 1 部分：LLM 评估函数
########################

def judge_replies(original_email: str, reply_a: str, reply_b: str) -> str:
    """
    对比 reply_a, reply_b，模型同时看到 original_email。
    返回:
    - "model1"  (表示 A 更好)
    - "model2"  (表示 B 更好)
    - "tie"     (表示 差不多)
    """
    client = OpenAI(api_key=os.environ["OPENAI_API_KEY"])

    system_prompt = """You are an unbiased judge. 
You will be provided with the original email and two email replies (Reply A and Reply B).
You must compare them based on the following two criteria:
1. Naturalness
2. Succinctness

Output exactly one single token among these:
- "model1"
- "model2"
- "tie"
Do not explain your reasoning. Just return one word exactly: model1, model2, or tie.
"""

    user_prompt = f"""
Original Email:
{original_email}

Reply A:
{reply_a}

Reply B:
{reply_b}

Which reply is better?
"""

    try:
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",  # 或 "gpt-4" / 你自己部署的模型
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user",   "content": user_prompt},
            ],
            temperature=0.0
        )
        result = response.choices[0].message.content.strip().lower()
        if result in ["model1", "model2", "tie"]:
            return result
        else:
            return "tie"
    except Exception as e:
        print("Error calling LLM:", e)
        return "tie"


########################
# 第 2 部分：Excel 数据读取 & 对比
########################

def compare_two_columns(
    workbook_path: str,
    col_a: int,
    col_b: int,
    skip_header: bool = True,
    judge_func=judge_replies,
    log_path: str = None,
    label: str = None
) -> dict:
    """
    从 Excel 中读取:
      - 第 1 列: original_email
      - 第 col_a 列: Reply A
      - 第 col_b 列: Reply B

    调用 judge_func(original_email, reply_a, reply_b) 进行对比，
    统计 modelA胜 / modelB胜 / tie 的次数。
    并可选：将每一条对比写入 log 文件 (CSV)。
    """
    wb = load_workbook(workbook_path)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if skip_header:
        rows = rows[1:]

    score_modelA = 0
    score_modelB = 0
    score_tie = 0
    total_count = 0

    logs = []

    for idx, row in enumerate(rows):
        # 第 1 列是原始邮件
        original_email = row[0] if len(row) >= 1 else None
        reply_a = row[col_a - 1] if len(row) >= col_a else None
        reply_b = row[col_b - 1] if len(row) >= col_b else None

        # 如果其中任何一个为空，则跳过
        if not original_email or not reply_a or not reply_b:
            continue

        total_count += 1

        # 记录对比开始时间
        start_time = time.time()
        result_raw = judge_func(original_email, reply_a, reply_b)
        elapsed = time.time() - start_time  # 对比花费的秒数

        # 将原始结果 "model1"/"model2"/"tie" 映射到日志表述
        if result_raw == "model1":
            result_str = "modelA胜"
            score_modelA += 1
        elif result_raw == "model2":
            result_str = "modelB胜"
            score_modelB += 1
        else:
            result_str = "tie"
            score_tie += 1

        # 添加到日志中
        logs.append({
            "Index": idx + 2 if skip_header else idx + 1,
            "Original Email": original_email,
            "Model A": reply_a,
            "Model B": reply_b,
            "Compare Result": result_str,
            "Time (sec)": f"{elapsed:.2f}"
        })

    # 如果指定了 log_path 和标签，就写入 CSV
    if log_path and label:
        if not os.path.exists(log_path):
            os.makedirs(log_path)
        filename = f"{log_path}/comparison_log_{label.replace(' ', '_')}.csv"
        with open(filename, mode="w", encoding="utf-8-sig", newline="") as f:
            # 这里多写一列 Original Email
            fieldnames = ["Index", "Original Email", "Model A", "Model B", "Compare Result", "Time (sec)"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(logs)
        print(f"🔍 Log saved to {filename}")

    return {
        "modelA": score_modelA,
        "modelB": score_modelB,
        "tie": score_tie,
        "total": total_count,
    }


########################
# 第 3 部分：可视化 - 画堆叠条形图
########################

def create_stacked_bar_chart(results_list: list, title: str = "Human Evaluation"):
    """
    给定若干组对比结果，画出堆叠条形图(横向柱状图)：
      - modelA胜 (底层)
      - tie (中间)
      - modelB胜 (上层)
    百分比堆叠。
    """
    labels = [r["label"] for r in results_list]

    modelA_pct = []
    tie_pct = []
    modelB_pct = []

    for r in results_list:
        total = r["total"]
        if total == 0:
            modelA_pct.append(0)
            tie_pct.append(0)
            modelB_pct.append(0)
        else:
            modelA_pct.append(r["modelA"] / total * 100)
            tie_pct.append(r["tie"] / total * 100)
            modelB_pct.append(r["modelB"] / total * 100)

    x_positions = range(len(labels))
    fig, ax = plt.subplots(figsize=(8, 4))

    # 底层：modelA胜
    barA = ax.barh(x_positions, modelA_pct, label='Win (ModelA)')
    # 中层：tie
    barT = ax.barh(x_positions, tie_pct, left=modelA_pct, label='Tie')
    # 上层：modelB胜
    left_for_B = [a + t for a, t in zip(modelA_pct, tie_pct)]
    barB = ax.barh(x_positions, modelB_pct, left=left_for_B, label='Win (ModelB)')

    def add_labels(bar_obj, base_list):
        for rect, base_val in zip(bar_obj, base_list):
            width = rect.get_width()
            if width > 0:
                x_pos = base_val + width / 2
                y_pos = rect.get_y() + rect.get_height() / 2
                ax.text(x_pos, y_pos, f"{width:.1f}%", ha='center', va='center', color='white', fontsize=9)

    # 在每段条形上标出百分比
    add_labels(barA, [0]*len(modelA_pct))
    add_labels(barT, modelA_pct)
    add_labels(barB, left_for_B)

    ax.set_yticks(list(x_positions))
    ax.set_yticklabels(labels)
    ax.set_xlabel("% win rate")
    ax.set_xlim([0, 100])
    ax.set_title(title)
    ax.invert_yaxis()  # 让第一组排在最上面
    ax.legend(loc="lower right")
    plt.tight_layout()
    plt.show()


########################
# 第 4 部分：主函数
########################

def main():
    workbook_path = "enron/enron_output.xlsx"
    
    # 你想要对比的四组
    # 三元组: (可视化/日志的label, colA, colB)
    pairs_to_compare = [
        ("gemma3:1b vs GPT-4o", 2, 6),
        ("gemma:2b vs GPT-4o", 3, 6),
        ("llama3.2:3b vs GPT-4o", 4, 6),
        ("mistral vs GPT-4o", 5, 6),
    ]
    
    results_for_chart = []
    
    for pair_label, colA, colB in pairs_to_compare:
        comparison = compare_two_columns(
            workbook_path=workbook_path,
            col_a=colA,
            col_b=colB,
            skip_header=True,
            judge_func=judge_replies,
            log_path="dataset/complex_gpt3",  # 日志文件夹
            label=pair_label
        )
        
        results_for_chart.append({
            "label": pair_label,
            "modelA": comparison["modelA"],
            "modelB": comparison["modelB"],
            "tie": comparison["tie"],
            "total": comparison["total"],
        })

    # 画一个堆叠条形图
    create_stacked_bar_chart(results_for_chart, title="Pairwise Comparison Results")


if __name__ == "__main__":
    main()
